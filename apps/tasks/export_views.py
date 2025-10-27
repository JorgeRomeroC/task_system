import csv
from io import BytesIO
from django.http import HttpResponse
from django.db.models import Count, Q
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import Task
from datetime import datetime


def export_tasks_csv(request):
    """Exporta las tareas a formato CSV."""

    # Crear la respuesta HTTP con el tipo de contenido CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="tareas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    # Agregar BOM para Excel
    response.write('\ufeff')

    writer = csv.writer(response)

    # Escribir encabezados
    writer.writerow([
        'ID',
        'Título',
        'Descripción',
        'Estado',
        'Fecha de Creación',
        'Última Actualización'
    ])

    # Escribir datos de las tareas
    tasks = Task.objects.all().order_by('-created_at')
    for task in tasks:
        writer.writerow([
            task.id,
            task.title,
            task.description,
            'Completada' if task.completed else 'Pendiente',
            task.created_at.strftime('%d/%m/%Y %H:%M'),
            task.updated_at.strftime('%d/%m/%Y %H:%M'),
        ])

    return response


def export_tasks_excel(request):
    """Exporta las tareas a formato Excel (.xlsx)."""

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tareas"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados
    headers = ['ID', 'Título', 'Descripción', 'Estado', 'Fecha de Creación', 'Última Actualización']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Datos
    tasks = Task.objects.all().order_by('-created_at')
    for row_num, task in enumerate(tasks, 2):
        ws.cell(row=row_num, column=1).value = task.id
        ws.cell(row=row_num, column=2).value = task.title
        ws.cell(row=row_num, column=3).value = task.description
        ws.cell(row=row_num, column=4).value = 'Completada' if task.completed else 'Pendiente'
        ws.cell(row=row_num, column=5).value = task.created_at.strftime('%d/%m/%Y %H:%M')
        ws.cell(row=row_num, column=6).value = task.updated_at.strftime('%d/%m/%Y %H:%M')

        # Aplicar bordes
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = border
            ws.cell(row=row_num, column=col).alignment = Alignment(vertical="center")

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    # Hoja de estadísticas
    ws_stats = wb.create_sheet("Estadísticas")

    # Calcular estadísticas
    total = Task.objects.count()
    completed = Task.objects.filter(completed=True).count()
    pending = Task.objects.filter(completed=False).count()
    percentage = (completed / total * 100) if total > 0 else 0

    # Encabezado de estadísticas
    stats_headers = ['Métrica', 'Valor']
    for col_num, header in enumerate(stats_headers, 1):
        cell = ws_stats.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Datos de estadísticas
    stats_data = [
        ['Total de Tareas', total],
        ['Tareas Completadas', completed],
        ['Tareas Pendientes', pending],
        ['Porcentaje Completado', f'{percentage:.2f}%'],
    ]

    for row_num, (metric, value) in enumerate(stats_data, 2):
        ws_stats.cell(row=row_num, column=1).value = metric
        ws_stats.cell(row=row_num, column=2).value = value

        for col in range(1, 3):
            ws_stats.cell(row=row_num, column=col).border = border
            ws_stats.cell(row=row_num, column=col).alignment = Alignment(vertical="center")

    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 15

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Crear respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="tareas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    return response


def export_tasks_pdf(request):
    """Exporta las tareas a formato PDF."""

    # Crear el objeto BytesIO
    buffer = BytesIO()

    # Crear el PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
        spaceBefore=12
    )

    # Título
    title = Paragraph("Reporte de Tareas", title_style)
    elements.append(title)

    # Fecha del reporte
    date_text = Paragraph(
        f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    )
    elements.append(date_text)
    elements.append(Spacer(1, 20))

    # Estadísticas
    stats_title = Paragraph("Estadísticas Generales", heading_style)
    elements.append(stats_title)

    total = Task.objects.count()
    completed = Task.objects.filter(completed=True).count()
    pending = Task.objects.filter(completed=False).count()
    percentage = (completed / total * 100) if total > 0 else 0

    stats_data = [
        ['Métrica', 'Valor'],
        ['Total de Tareas', str(total)],
        ['Tareas Completadas', str(completed)],
        ['Tareas Pendientes', str(pending)],
        ['Porcentaje Completado', f'{percentage:.2f}%'],
    ]

    stats_table = Table(stats_data, colWidths=[3 * inch, 2 * inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))

    elements.append(stats_table)
    elements.append(Spacer(1, 30))

    # Lista de tareas
    tasks_title = Paragraph("Detalle de Tareas", heading_style)
    elements.append(tasks_title)

    tasks = Task.objects.all().order_by('-created_at')

    if tasks.exists():
        tasks_data = [['ID', 'Título', 'Estado', 'Fecha Creación']]

        for task in tasks:
            tasks_data.append([
                str(task.id),
                task.title[:40] + '...' if len(task.title) > 40 else task.title,
                '✓ Completada' if task.completed else '○ Pendiente',
                task.created_at.strftime('%d/%m/%Y'),
            ])

        tasks_table = Table(tasks_data, colWidths=[0.5 * inch, 3.5 * inch, 1.5 * inch, 1.5 * inch])
        tasks_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(tasks_table)
    else:
        no_tasks = Paragraph("No hay tareas para mostrar.", styles['Normal'])
        elements.append(no_tasks)

    # Construir PDF
    doc.build(elements)

    # Obtener el valor del buffer
    pdf = buffer.getvalue()
    buffer.close()

    # Crear respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="tareas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response.write(pdf)

    return response